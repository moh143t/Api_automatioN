from openpyxl import load_workbook
import os
import datetime


def readConfigValue(row,column,filepath):

    wb = load_workbook (filepath)
    ws = wb['Config']
    Value = ws.cell(row=row, column=column).value
    return Value

def writeConfigValue(row,column,valueToWrite,filepath):
    wb = load_workbook (filepath)
    ws = wb['Config']
    ws.cell(row=row,column=column).value = valueToWrite
    writedone=wb.save(filepath)
    return writedone

def readRowCount(filepath):

    wb = load_workbook (filepath)
    ws = wb['API']
    row_count=ws.max_row
    return row_count


def readCellValue(row,column,filepath):

    wb = load_workbook (filepath)
    ws = wb['API']
    Value = ws.cell(row=row, column=column).value
    return Value


def writeCellValue(row,column,valueToWrite,filepath):
    wb = load_workbook (filepath)
    ws = wb['API']
    ws.cell(row=row,column=column).value = valueToWrite
    writedone=wb.save(filepath)
    return writedone

def date_execution():
    d=(datetime.datetime.now())
    d1=(d.strftime("%A"))+"   "+(d.strftime ("%c"))
    return d1

def fail_count(filepath):
        failed=0
        for i in range (2, (readRowCount (filepath) + 1)):
            if (readCellValue (i, 8, filepath) == 'Fails'):
                failed+=1
        return failed


def pass_count(filepath):
    passed = 0
    for i in range (2, (readRowCount (filepath) + 1)):
        if (readCellValue (i, 8, filepath) == 'Pass'):
            passed += 1
    return passed



def Files_path_list(directory_path):
    arr = os.listdir (directory_path)
    files=[]
    for a in arr:
        filePath = os.path.abspath (directory_path)
        filePathWithSlash = filePath + "\\"
        filenameWithPath = os.path.join (filePathWithSlash, a)
        files.append (filenameWithPath)
    return files


